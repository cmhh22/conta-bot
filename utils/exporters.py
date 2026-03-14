"""
Módulo para generar reportes en formato PDF y Excel.
"""
import logging
import os
from datetime import datetime
from typing import List, Dict, Any, Optional
from pathlib import Path

logger = logging.getLogger(__name__)

# Intentar importar las librerías necesarias
try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False
    logger.warning("reportlab no está instalado. Las exportaciones a PDF no estarán disponibles.")

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    logger.warning("openpyxl no está instalado. Las exportaciones a Excel no estarán disponibles.")


def _ensure_export_dir() -> Path:
    """Asegura que existe el directorio de exportaciones."""
    export_dir = Path("exports")
    export_dir.mkdir(exist_ok=True)
    return export_dir


def export_movimientos_pdf(movimientos: List[Dict[str, Any]], filename: Optional[str] = None) -> str:
    """
    Exporta movimientos a un archivo PDF.
    
    Args:
        movimientos: Lista de diccionarios con los movimientos
        filename: Nombre del archivo (opcional)
    
    Returns:
        Ruta del archivo PDF generado
    
    Raises:
        ImportError: Si reportlab no está instalado
    """
    if not REPORTLAB_AVAILABLE:
        raise ImportError("reportlab no está instalado. Instala con: pip install reportlab")
    
    if not filename:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"movimientos_{timestamp}.pdf"
    
    export_dir = _ensure_export_dir()
    filepath = export_dir / filename
    
    # Crear documento PDF
    doc = SimpleDocTemplate(str(filepath), pagesize=A4)
    story = []
    styles = getSampleStyleSheet()
    
    # Estilos personalizados
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        textColor=colors.HexColor('#1a1a1a'),
        spaceAfter=30,
        alignment=TA_CENTER
    )
    
    header_style = ParagraphStyle(
        'CustomHeader',
        parent=styles['Normal'],
        fontSize=10,
        textColor=colors.white,
        alignment=TA_CENTER,
        fontName='Helvetica-Bold'
    )
    
    # Título
    story.append(Paragraph("REPORTE DE MOVIMIENTOS", title_style))
    story.append(Spacer(1, 0.2*inch))
    
    # Información del reporte
    fecha_reporte = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
    info_text = f"<b>Fecha de generación:</b> {fecha_reporte}<br/>"
    info_text += f"<b>Total de movimientos:</b> {len(movimientos)}"
    story.append(Paragraph(info_text, styles['Normal']))
    story.append(Spacer(1, 0.3*inch))
    
    if not movimientos:
        story.append(Paragraph("No hay movimientos para mostrar.", styles['Normal']))
    else:
        # Preparar datos para la tabla
        data = []
        
        # Encabezados
        headers = ['Fecha', 'Tipo', 'Monto', 'Moneda', 'Caja', 'Descripción']
        data.append(headers)
        
        # Datos
        for mov in movimientos:
            fecha = mov.get('fecha', 'N/A')
            if isinstance(fecha, str) and ' ' in fecha:
                fecha = fecha.split()[0]  # Solo la fecha
            
            tipo = mov.get('tipo', 'N/A').upper()
            monto = mov.get('monto', 0)
            moneda = mov.get('moneda', 'N/A').upper()
            caja_nombre = mov.get('caja_nombre', mov.get('caja', 'N/A'))
            descripcion = mov.get('descripcion', '')[:50]  # Limitar longitud
            
            row = [
                fecha,
                tipo,
                f"{monto:,.2f}",
                moneda,
                caja_nombre,
                descripcion
            ]
            data.append(row)
        
        # Crear tabla
        table = Table(data, colWidths=[1*inch, 0.8*inch, 1*inch, 0.6*inch, 1*inch, 2*inch])
        
        # Estilo de la tabla
        table.setStyle(TableStyle([
            # Encabezados
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('TOPPADDING', (0, 0), (-1, 0), 12),
            
            # Filas alternadas
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F2F2F2')]),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))
        
        story.append(table)
    
    # Construir PDF
    doc.build(story)
    logger.info(f"PDF generado: {filepath}")
    return str(filepath)


def export_movimientos_excel(movimientos: List[Dict[str, Any]], filename: Optional[str] = None) -> str:
    """
    Exporta movimientos a un archivo Excel.
    
    Args:
        movimientos: Lista de diccionarios con los movimientos
        filename: Nombre del archivo (opcional)
    
    Returns:
        Ruta del archivo Excel generado
    
    Raises:
        ImportError: Si openpyxl no está instalado
    """
    if not OPENPYXL_AVAILABLE:
        raise ImportError("openpyxl no está instalado. Instala con: pip install openpyxl")
    
    if not filename:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"movimientos_{timestamp}.xlsx"
    
    export_dir = _ensure_export_dir()
    filepath = export_dir / filename
    
    # Crear workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Movimientos"
    
    # Estilos
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_align = Alignment(horizontal='center', vertical='center')
    
    # Encabezados
    headers = ['ID', 'Fecha', 'Tipo', 'Monto', 'Moneda', 'Caja', 'Usuario ID', 'Descripción']
    ws.append(headers)
    
    # Aplicar estilo a encabezados
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = border
    
    # Datos
    for mov in movimientos:
        fecha = mov.get('fecha', 'N/A')
        if isinstance(fecha, str) and ' ' in fecha:
            fecha = fecha.split()[0]  # Solo la fecha
        
        row = [
            mov.get('id', ''),
            fecha,
            mov.get('tipo', 'N/A').upper(),
            mov.get('monto', 0),
            mov.get('moneda', 'N/A').upper(),
            mov.get('caja_nombre', mov.get('caja', 'N/A')),
            mov.get('user_id', ''),
            mov.get('descripcion', '')
        ]
        ws.append(row)
        
        # Aplicar bordes a la fila
        for col_num in range(1, len(headers) + 1):
            cell = ws.cell(row=ws.max_row, column=col_num)
            cell.border = border
    
    # Ajustar ancho de columnas
    column_widths = {
        'A': 8,   # ID
        'B': 12,  # Fecha
        'C': 12,  # Tipo
        'D': 12,  # Monto
        'E': 10,  # Moneda
        'F': 15,  # Caja
        'G': 12,  # Usuario ID
        'H': 40,  # Descripción
    }
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    # Congelar primera fila
    ws.freeze_panes = 'A2'
    
    # Guardar archivo
    wb.save(str(filepath))
    logger.info(f"Excel generado: {filepath}")
    return str(filepath)


def export_inventario_pdf(productos: List[Dict[str, Any]], filename: Optional[str] = None) -> str:
    """
    Exporta inventario de productos a un archivo PDF.
    
    Args:
        productos: Lista de diccionarios con los productos
        filename: Nombre del archivo (opcional)
    
    Returns:
        Ruta del archivo PDF generado
    """
    if not REPORTLAB_AVAILABLE:
        raise ImportError("reportlab no está instalado. Instala con: pip install reportlab")
    
    if not filename:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"inventario_{timestamp}.pdf"
    
    export_dir = _ensure_export_dir()
    filepath = export_dir / filename
    
    doc = SimpleDocTemplate(str(filepath), pagesize=A4)
    story = []
    styles = getSampleStyleSheet()
    
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        textColor=colors.HexColor('#1a1a1a'),
        spaceAfter=30,
        alignment=TA_CENTER
    )
    
    # Título
    story.append(Paragraph("REPORTE DE INVENTARIO", title_style))
    story.append(Spacer(1, 0.2*inch))
    
    fecha_reporte = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
    info_text = f"<b>Fecha de generación:</b> {fecha_reporte}<br/>"
    info_text += f"<b>Total de productos:</b> {len(productos)}"
    story.append(Paragraph(info_text, styles['Normal']))
    story.append(Spacer(1, 0.3*inch))
    
    if not productos:
        story.append(Paragraph("No hay productos en el inventario.", styles['Normal']))
    else:
        data = []
        headers = ['Código', 'Nombre', 'Stock', 'Costo Unit.', 'Moneda', 'Precio Venta']
        data.append(headers)
        
        for prod in productos:
            row = [
                prod.get('codigo', 'N/A'),
                prod.get('nombre', 'N/A')[:30],
                f"{prod.get('stock', 0):.2f}",
                f"{prod.get('costo_unitario', 0):,.2f}",
                prod.get('moneda_costo', 'N/A').upper(),
                f"{prod.get('precio_venta', 0):,.2f}" if prod.get('precio_venta') else 'N/A'
            ]
            data.append(row)
        
        table = Table(data, colWidths=[1*inch, 2.5*inch, 0.8*inch, 1*inch, 0.7*inch, 1*inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F2F2F2')]),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))
        story.append(table)
    
    doc.build(story)
    logger.info(f"PDF de inventario generado: {filepath}")
    return str(filepath)


def export_inventario_excel(productos: List[Dict[str, Any]], filename: Optional[str] = None) -> str:
    """
    Exporta inventario de productos a un archivo Excel.
    
    Args:
        productos: Lista de diccionarios con los productos
        filename: Nombre del archivo (opcional)
    
    Returns:
        Ruta del archivo Excel generado
    """
    if not OPENPYXL_AVAILABLE:
        raise ImportError("openpyxl no está instalado. Instala con: pip install openpyxl")
    
    if not filename:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"inventario_{timestamp}.xlsx"
    
    export_dir = _ensure_export_dir()
    filepath = export_dir / filename
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Inventario"
    
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    headers = ['Código', 'Nombre', 'Stock', 'Costo Unitario', 'Moneda Costo', 'Precio Venta']
    ws.append(headers)
    
    for col_num in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
    
    for prod in productos:
        row = [
            prod.get('codigo', ''),
            prod.get('nombre', ''),
            prod.get('stock', 0),
            prod.get('costo_unitario', 0),
            prod.get('moneda_costo', ''),
            prod.get('precio_venta', '') if prod.get('precio_venta') else None
        ]
        ws.append(row)
        
        for col_num in range(1, len(headers) + 1):
            cell = ws.cell(row=ws.max_row, column=col_num)
            cell.border = border
    
    column_widths = {'A': 15, 'B': 30, 'C': 12, 'D': 15, 'E': 12, 'F': 15}
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    ws.freeze_panes = 'A2'
    wb.save(str(filepath))
    logger.info(f"Excel de inventario generado: {filepath}")
    return str(filepath)


def export_deudas_pdf(deudas: Dict[str, List[Dict[str, Any]]], filename: Optional[str] = None) -> str:
    """
    Exporta deudas a un archivo PDF.
    
    Args:
        deudas: Diccionario con 'por_pagar' y 'por_cobrar'
        filename: Nombre del archivo (opcional)
    
    Returns:
        Ruta del archivo PDF generado
    """
    if not REPORTLAB_AVAILABLE:
        raise ImportError("reportlab no está instalado. Instala con: pip install reportlab")
    
    if not filename:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"deudas_{timestamp}.pdf"
    
    export_dir = _ensure_export_dir()
    filepath = export_dir / filename
    
    doc = SimpleDocTemplate(str(filepath), pagesize=A4)
    story = []
    styles = getSampleStyleSheet()
    
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        textColor=colors.HexColor('#1a1a1a'),
        spaceAfter=30,
        alignment=TA_CENTER
    )
    
    story.append(Paragraph("REPORTE DE DEUDAS", title_style))
    story.append(Spacer(1, 0.2*inch))
    
    fecha_reporte = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
    story.append(Paragraph(f"<b>Fecha de generación:</b> {fecha_reporte}", styles['Normal']))
    story.append(Spacer(1, 0.3*inch))
    
    # Cuentas por pagar
    story.append(Paragraph("<b>CUENTAS POR PAGAR (Proveedores)</b>", styles['Heading2']))
    story.append(Spacer(1, 0.1*inch))
    
    if deudas.get('por_pagar'):
        data = [['Proveedor', 'Monto', 'Moneda']]
        for deuda in deudas['por_pagar']:
            data.append([
                deuda.get('actor_id', 'N/A'),
                f"{deuda.get('monto', 0):,.2f}",
                deuda.get('moneda', 'N/A').upper()
            ])
        
        table = Table(data, colWidths=[3*inch, 1.5*inch, 1*inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#DC3545')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#FFE5E5')]),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ]))
        story.append(table)
    else:
        story.append(Paragraph("<i>No hay deudas por pagar.</i>", styles['Normal']))
    
    story.append(Spacer(1, 0.3*inch))
    
    # Cuentas por cobrar
    story.append(Paragraph("<b>CUENTAS POR COBRAR (Vendedores)</b>", styles['Heading2']))
    story.append(Spacer(1, 0.1*inch))
    
    if deudas.get('por_cobrar'):
        data = [['Vendedor', 'Monto', 'Moneda']]
        for deuda in deudas['por_cobrar']:
            data.append([
                deuda.get('actor_id', 'N/A'),
                f"{deuda.get('monto', 0):,.2f}",
                deuda.get('moneda', 'N/A').upper()
            ])
        
        table = Table(data, colWidths=[3*inch, 1.5*inch, 1*inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#28A745')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#E5FFE5')]),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ]))
        story.append(table)
    else:
        story.append(Paragraph("<i>No hay deudas por cobrar.</i>", styles['Normal']))
    
    doc.build(story)
    logger.info(f"PDF de deudas generado: {filepath}")
    return str(filepath)


def export_deudas_excel(deudas: Dict[str, List[Dict[str, Any]]], filename: Optional[str] = None) -> str:
    """
    Exporta deudas a un archivo Excel.
    
    Args:
        deudas: Diccionario con 'por_pagar' y 'por_cobrar'
        filename: Nombre del archivo (opcional)
    
    Returns:
        Ruta del archivo Excel generado
    """
    if not OPENPYXL_AVAILABLE:
        raise ImportError("openpyxl no está instalado. Instala con: pip install openpyxl")
    
    if not filename:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"deudas_{timestamp}.xlsx"
    
    export_dir = _ensure_export_dir()
    filepath = export_dir / filename
    
    wb = Workbook()
    
    # Hoja de cuentas por pagar
    ws_pagar = wb.active
    ws_pagar.title = "Por Pagar"
    
    header_fill_red = PatternFill(start_color="DC3545", end_color="DC3545", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    headers = ['Proveedor', 'Monto', 'Moneda']
    ws_pagar.append(headers)
    
    for col_num in range(1, len(headers) + 1):
        cell = ws_pagar.cell(row=1, column=col_num)
        cell.fill = header_fill_red
        cell.font = header_font
        cell.border = border
    
    if deudas.get('por_pagar'):
        for deuda in deudas['por_pagar']:
            row = [
                deuda.get('actor_id', ''),
                deuda.get('monto', 0),
                deuda.get('moneda', '').upper()
            ]
            ws_pagar.append(row)
            for col_num in range(1, len(headers) + 1):
                ws_pagar.cell(row=ws_pagar.max_row, column=col_num).border = border
    
    # Ajustar columnas
    ws_pagar.column_dimensions['A'].width = 30
    ws_pagar.column_dimensions['B'].width = 15
    ws_pagar.column_dimensions['C'].width = 10
    
    # Hoja de cuentas por cobrar
    ws_cobrar = wb.create_sheet("Por Cobrar")
    
    header_fill_green = PatternFill(start_color="28A745", end_color="28A745", fill_type="solid")
    ws_cobrar.append(headers)
    
    for col_num in range(1, len(headers) + 1):
        cell = ws_cobrar.cell(row=1, column=col_num)
        cell.fill = header_fill_green
        cell.font = header_font
        cell.border = border
    
    if deudas.get('por_cobrar'):
        for deuda in deudas['por_cobrar']:
            row = [
                deuda.get('actor_id', ''),
                deuda.get('monto', 0),
                deuda.get('moneda', '').upper()
            ]
            ws_cobrar.append(row)
            for col_num in range(1, len(headers) + 1):
                ws_cobrar.cell(row=ws_cobrar.max_row, column=col_num).border = border
    
    ws_cobrar.column_dimensions['A'].width = 30
    ws_cobrar.column_dimensions['B'].width = 15
    ws_cobrar.column_dimensions['C'].width = 10
    
    wb.save(str(filepath))
    logger.info(f"Excel de deudas generado: {filepath}")
    return str(filepath)


def export_cajas_externas_pdf(transferencias: List[Dict[str, Any]], caja_externa: Dict[str, Any], filename: Optional[str] = None) -> str:
    """
    Exporta transferencias de una caja externa a PDF.
    
    Args:
        transferencias: Lista de transferencias
        caja_externa: Información de la caja externa
        filename: Nombre del archivo (opcional)
    
    Returns:
        Ruta del archivo PDF generado
    """
    if not REPORTLAB_AVAILABLE:
        raise ImportError("reportlab no está instalado. Instala con: pip install reportlab")
    
    if not filename:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"caja_externa_{caja_externa.get('nombre', 'unknown')}_{timestamp}.pdf"
    
    export_dir = _ensure_export_dir()
    filepath = export_dir / filename
    
    doc = SimpleDocTemplate(str(filepath), pagesize=A4)
    story = []
    styles = getSampleStyleSheet()
    
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        textColor=colors.HexColor('#1a1a1a'),
        spaceAfter=30,
        alignment=TA_CENTER
    )
    
    story.append(Paragraph(f"REPORTE DE CAJA EXTERNA: {caja_externa.get('nombre', 'N/A')}", title_style))
    story.append(Spacer(1, 0.2*inch))
    
    info_text = f"<b>Ubicación:</b> {caja_externa.get('ubicacion', 'N/A')}<br/>"
    info_text += f"<b>Porcentaje Envío:</b> {caja_externa.get('porcentaje_envio', 0):.2f}%<br/>"
    info_text += f"<b>Total Transferencias:</b> {len(transferencias)}<br/>"
    info_text += f"<b>Fecha de generación:</b> {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"
    story.append(Paragraph(info_text, styles['Normal']))
    story.append(Spacer(1, 0.3*inch))
    
    if transferencias:
        data = [['Fecha', 'Producto', 'Monto', 'Moneda', 'Envío', 'Recibido', 'Caja Origen']]
        
        for transf in transferencias:
            fecha = transf.get('fecha', 'N/A')
            if isinstance(fecha, str) and ' ' in fecha:
                fecha = fecha.split()[0]
            
            data.append([
                fecha,
                f"{transf.get('producto_codigo', 'N/A')} - {transf.get('producto_nombre', 'N/A')[:20]}",
                f"{transf.get('monto', 0):,.2f}",
                transf.get('moneda', 'N/A').upper(),
                f"{transf.get('monto_envio', 0):,.2f}",
                f"{transf.get('monto_recibido', 0):,.2f}",
                transf.get('caja_origen_nombre', 'N/A')
            ])
        
        table = Table(data, colWidths=[0.9*inch, 1.8*inch, 0.9*inch, 0.7*inch, 0.9*inch, 0.9*inch, 1.2*inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 8),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F2F2F2')]),
            ('FONTSIZE', (0, 1), (-1, -1), 7),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))
        story.append(table)
        
        # Resumen por moneda
        story.append(Spacer(1, 0.3*inch))
        story.append(Paragraph("<b>RESUMEN POR MONEDA</b>", styles['Heading3']))
        
        totales_por_moneda = {}
        for transf in transferencias:
            moneda = transf.get('moneda', 'N/A')
            if moneda not in totales_por_moneda:
                totales_por_moneda[moneda] = {'total': 0, 'envio': 0, 'recibido': 0}
            totales_por_moneda[moneda]['total'] += transf.get('monto', 0)
            totales_por_moneda[moneda]['envio'] += transf.get('monto_envio', 0)
            totales_por_moneda[moneda]['recibido'] += transf.get('monto_recibido', 0)
        
        resumen_data = [['Moneda', 'Total Transferido', 'Total Envío', 'Total Recibido']]
        for moneda, totales in totales_por_moneda.items():
            resumen_data.append([
                moneda.upper(),
                f"{totales['total']:,.2f}",
                f"{totales['envio']:,.2f}",
                f"{totales['recibido']:,.2f}"
            ])
        
        resumen_table = Table(resumen_data, colWidths=[1*inch, 1.5*inch, 1.5*inch, 1.5*inch])
        resumen_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#28A745')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ]))
        story.append(resumen_table)
    else:
        story.append(Paragraph("<i>No hay transferencias registradas.</i>", styles['Normal']))
    
    doc.build(story)
    logger.info(f"PDF de caja externa generado: {filepath}")
    return str(filepath)


def export_cajas_externas_excel(transferencias: List[Dict[str, Any]], caja_externa: Dict[str, Any], filename: Optional[str] = None) -> str:
    """
    Exporta transferencias de una caja externa a Excel.
    
    Args:
        transferencias: Lista de transferencias
        caja_externa: Información de la caja externa
        filename: Nombre del archivo (opcional)
    
    Returns:
        Ruta del archivo Excel generado
    """
    if not OPENPYXL_AVAILABLE:
        raise ImportError("openpyxl no está instalado. Instala con: pip install openpyxl")
    
    if not filename:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"caja_externa_{caja_externa.get('nombre', 'unknown')}_{timestamp}.xlsx"
    
    export_dir = _ensure_export_dir()
    filepath = export_dir / filename
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Transferencias"
    
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Información de la caja externa
    ws.append(['Caja Externa:', caja_externa.get('nombre', 'N/A')])
    ws.append(['Ubicación:', caja_externa.get('ubicacion', 'N/A')])
    ws.append(['Porcentaje Envío:', f"{caja_externa.get('porcentaje_envio', 0):.2f}%"])
    ws.append(['Fecha de generación:', datetime.now().strftime('%d/%m/%Y %H:%M:%S')])
    ws.append([])
    
    headers = ['Fecha', 'Producto Código', 'Producto Nombre', 'Monto', 'Moneda', 'Envío', 'Recibido', 'Caja Origen']
    ws.append(headers)
    
    for col_num in range(1, len(headers) + 1):
        cell = ws.cell(row=6, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
    
    for transf in transferencias:
        fecha = transf.get('fecha', 'N/A')
        if isinstance(fecha, str) and ' ' in fecha:
            fecha = fecha.split()[0]
        
        row = [
            fecha,
            transf.get('producto_codigo', ''),
            transf.get('producto_nombre', ''),
            transf.get('monto', 0),
            transf.get('moneda', '').upper(),
            transf.get('monto_envio', 0),
            transf.get('monto_recibido', 0),
            transf.get('caja_origen_nombre', '')
        ]
        ws.append(row)
        
        for col_num in range(1, len(headers) + 1):
            cell = ws.cell(row=ws.max_row, column=col_num)
            cell.border = border
    
    column_widths = {'A': 12, 'B': 15, 'C': 25, 'D': 12, 'E': 10, 'F': 12, 'G': 12, 'H': 20}
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    ws.freeze_panes = 'A7'
    wb.save(str(filepath))
    logger.info(f"Excel de caja externa generado: {filepath}")
    return str(filepath)

